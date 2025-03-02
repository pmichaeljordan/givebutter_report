#!/usr/bin/env python3
import os
import json
import datetime
import requests
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Google Drive imports
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow

# Base URL for the API
base_url = "https://api.givebutter.com/v1/"

# Headers with authorization and accepting JSON response
headers = {
    "accept": "application/json",
    "Authorization": f"Bearer {os.getenv('GIVEBUTTER_API_TOKEN')}",
}
auth_token = os.getenv("GIVEBUTTER_API_TOKEN")

# Google Drive folder ID (replace with your actual folder ID)
drive_folder_id = "1aFkyJD_Qtto8ZOVZ8Q6FZLS5rBF2ERcw"

# Define the Drive API scopes
DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive.file']

def get_credentials(scopes, token_file='drive_token.json', credentials_file='credentials.json'):
    """
    Obtains OAuth credentials using the local token if available,
    or runs the OAuth flow if not.
    """
    creds = None
    if os.path.exists(token_file):
        creds = Credentials.from_authorized_user_file(token_file, scopes)
    if not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file(credentials_file, scopes)
        creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(token_file, 'w') as token:
            token.write(creds.to_json())
    return creds

def get_campaign(auth_token):
    """
    Function to list campaigns and allow the user to select one.
    """
    url = "https://api.givebutter.com/v1/campaigns"
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        try:
            campaign_data = response.json()
            campaigns = campaign_data.get("data", [])
            if not campaigns:
                print("No campaign data found in response.")
                return None

            print("Available Campaigns:")
            for index, campaign in enumerate(campaigns):
                name = campaign.get("name", "No name")
                camp_id = campaign.get("id", "N/A")
                print(f"{index + 1}. {name} (ID: {camp_id})")

            selected = input("Enter the number of the campaign you want to use: ")
            try:
                selected_index = int(selected) - 1
                if 0 <= selected_index < len(campaigns):
                    campaign_id = campaigns[selected_index]["id"]
                    print("Selected Campaign ID:", campaign_id)
                    return campaign_id
                else:
                    print("Invalid selection. Defaulting to the first campaign.")
                    return campaigns[0]["id"]
            except ValueError:
                print("Invalid input. Defaulting to the first campaign.")
                return campaigns[0]["id"]
        except json.JSONDecodeError:
            print("Error parsing response as JSON. Response text:", response.text)
    else:
        print(f"Failed to retrieve campaigns: Status code {response.status_code}")
        print("Response:", response.text)
    return None

def get_campaign_members(campaign_id):
    """
    Function to retrieve members of a specific campaign using its ID.
    """
    page = 1
    all_members = []
    while True:
        members_url = f"https://api.givebutter.com/v1/campaigns/{campaign_id}/members?page={page}"
        response = requests.get(members_url, headers=headers)

        if response.status_code == 200:
            try:
                members_data = json.loads(response.text)
            except json.JSONDecodeError:
                return f"Failed to parse response as JSON: {response.text}"

            member_data = members_data["data"]
            all_members.extend(member_data)
            page += 1
            if members_data["meta"]["current_page"] <= members_data["meta"]["last_page"]:
                continue
            else:
                break
        else:
            return f"Failed to retrieve members: Status code {response.status_code}\nResponse: {response.text}"
    df2 = pd.DataFrame(all_members)
    return df2

def get_tickets():
    page = 1
    all_tickets = []
    while True:
        ticket_url = f"https://api.givebutter.com/v1/tickets?page={page}"
        response = requests.get(ticket_url, headers=headers)
        if response.status_code == 200:
            tickets_data = json.loads(response.text)
            ticket_data = tickets_data["data"]
            all_tickets.extend(ticket_data)
            page += 1
            if tickets_data["meta"]["current_page"] <= tickets_data["meta"]["last_page"]:
                continue
            else:
                break
        else:
            return f"Failed to retrieve tickets: Status code {response.status_code}\nResponse: {response.text}"
    return all_tickets

def format_data():
    """
    Generates the GiveButterReport.xlsx file for tickets that have titles starting with the current year.
    """
    df = pd.DataFrame(get_tickets())

    # Convert email addresses to lowercase
    df["email"] = df["email"].str.lower()

    # Filter tickets by title: keep only tickets where title starts with the current year.
    current_year_str = str(datetime.datetime.now().year)
    if "title" in df.columns:
        df = df[df["title"].str.startswith(current_year_str)]
    else:
        print("Warning: 'title' field not found in tickets. No filtering applied.")

    # Extract the desired columns
    columns = ["name", "first_name", "last_name", "email", "phone", "title", "price", "created_at"]
    data = df[columns].copy()

    # Convert 'created_at' to datetime and format it to 'YYYY-MM-DD'
    data["created_at"] = pd.to_datetime(data["created_at"]).dt.strftime("%Y-%m-%d")
    data = data.rename(
        columns={
            "name": "Name",
            "first_name": "First",
            "email": "Email",
            "phone": "Phone",
            "title": "Title",
            "price": "Price",
            "created_at": "Signup Date",
        }
    )

    # Group the data by title and get the count of each title
    title_counts = data.groupby("Title").size()

    wb = openpyxl.Workbook()
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    for title, count in title_counts.items():
        # Create a sheet name based on title, truncate to 31 characters
        sheet_name = title.rsplit("-", 1)[-1]
        sheet_name = sheet_name.replace("/", "_")[:31]
        ws = wb.create_sheet(sheet_name)
        ws.append(data.columns.tolist())
        filtered_data = data[data["Title"] == title]
        for index, row in filtered_data.iterrows():
            ws.append(row.tolist())

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    file_name = "GiveButterReport.xlsx"
    path = os.path.join(os.getcwd(), file_name)
    wb.save(path)
    print(f"Generated Excel file: {path}")

def fundraising(df2: pd.DataFrame, file_name: str = "Fundraising_Progress.xlsx"):
    """
    Generates the Fundraising_Progress.xlsx report.
    """
    df2 = df2.drop(columns=["id", "picture", "items", "url"])
    df2 = df2.rename(
        columns={
            "first_name": "First Name",
            "last_name": "Last Name",
            "display_name": "Display Name",
            "email": "Email",
            "phone": "Phone",
            "raised": "Raised",
            "goal": "Goal",
            "donors": "Donors",
        }
    )
    total_raised = df2["Raised"].sum()
    wb = openpyxl.Workbook()
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)
    ws = wb.create_sheet("Fundraising")
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws.append(r)
    ws.append(["", "", "", "", "Total Raised", total_raised, "", ""])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(file_name)
    print(f"Generated Excel file: {file_name}")
    return file_name

def upload_to_drive(file_path, folder_id):
    """
    Uploads the specified file to the given Google Drive folder using OAuth credentials.
    """
    creds = get_credentials(DRIVE_SCOPES, token_file='drive_token.json', credentials_file='credentials_account3.json')
    service = build('drive', 'v3', credentials=creds)
    file_metadata = {
        'name': os.path.basename(file_path),
        'parents': [folder_id]
    }
    media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    uploaded_file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    print(f"File uploaded to Google Drive. File ID: {uploaded_file.get('id')}")

# Main execution
campaign_id = get_campaign(auth_token)
if campaign_id:
    df2 = get_campaign_members(campaign_id)
    fundraising_file = fundraising(df2)
    format_data()
    # Upload the Fundraising_Progress.xlsx file to Google Drive
    upload_to_drive(fundraising_file, drive_folder_id)

